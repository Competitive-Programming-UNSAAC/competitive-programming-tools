import requests
import json
import openpyxl
import sys

from tabulate import tabulate


class Ranking:
    users = []
    totalUsers = 0
    ratingAverage = 0
    problemSolvedAverage = 0

    def __init__(self, users, totalUsers, ratingAverage, problemSolvedAverage):
        self.users = users
        self.totalUsers = totalUsers
        self.ratingAverage = ratingAverage
        self.problemSolvedAverage = problemSolvedAverage


class User:
    # Basic information
    id = ""
    name = ""
    handle = ""
    position = 0

    # Codeforces info
    rating = 0
    problems = 0

    # Points
    positionPoints = 0.0
    ratingPoints = 0.0
    problemPoints = 0.0
    totalPoints = 0.0

    def __init__(self, id, name, handle, position, rating, problems):
        self.id = id
        self.name = name
        self.handle = handle
        self.position = position
        self.rating = rating
        self.problems = problems


def numberOfProblemsByUser(handle):
    enlace = f"https://codeforces.com/api/user.status?handle={handle}"

    url = requests.get(enlace)
    text = url.text

    # convert to JSON
    data = json.loads(text)

    NroProblem = len(data["result"])

    problemSolved = set()

    for submissions in data["result"]:
        if submissions['verdict'] == "OK":
            key = ""

            if 'problemsetName' in submissions["problem"].keys():
                key += submissions["problem"]["problemsetName"]

            if 'contestId' in submissions["problem"].keys():
                key += str(submissions["problem"]["contestId"])

            if 'index' in submissions["problem"].keys():
                key += submissions["problem"]["index"]

            problemSolved.add(key)
    return len(problemSolved)


def ratingbyUser(handle):
    # create the link
    link = f"https://codeforces.com/api/user.info?handles={handle}"

    url = requests.get(link)
    text = url.text

    data = json.loads(text)

    rating = 0
    if 'rating' in data["result"][0].keys():
        rating = data["result"][0]["rating"]

    return rating


def readDataFromFile(filepath):
    print("Reading data from \"{0}\" file...".format(filepath))
    dataframe = openpyxl.load_workbook(filepath)
    data = dataframe.active
    print("File reading is completed!")
    return data


def getCompleteInfoByUser(data):
    # Format file
    # First rows column names
    id_col = 1
    name_col = 2
    position_col = 3
    handle_col = 4

    users = []

    print("Loading complete information...")
    for row in range(2, data.max_row + 1):
        id = data.cell(row, id_col).value
        name = data.cell(row, name_col).value
        position = data.cell(row, position_col).value
        handle = data.cell(row, handle_col).value

        if handle == "" or handle == None or id == "" or id == None:
            continue;

        print('Loading Codeforces info for \"{0}\" with handle \"{1}\"'.format(name, handle))
        codeforcesRating = ratingbyUser(handle)
        codeforcesNumberOfSolvedProblems = numberOfProblemsByUser(handle)

        print("Loading basic information for \"{0}\"".format(name))
        user = User(id, name, handle, position, codeforcesRating, codeforcesNumberOfSolvedProblems)
        users.append(user)

    print("Loading information is completed!")
    return users


def computeRanking(users, totalParticipants, ratingWeight, problemsWeight, positionWeight):
    print("Computing ranking...")

    # Computing Points
    totalUsers = len(users)
    ratingAverage = 0
    problemSolvedAverage = 0
    for user in users:
        ratingAverage += user.rating
        problemSolvedAverage += user.problems

    ratingAverage = ratingAverage/totalUsers
    problemSolvedAverage = problemSolvedAverage/totalUsers

    # Complete information
    for user in users:
        user.positionPoints = (totalParticipants - user.position) * positionWeight / totalParticipants;
        user.ratingPoints = max(ratingWeight, ratingWeight * user.rating / ratingAverage);
        user.problemsPoints = max(problemsWeight, problemsWeight * user.problems / problemSolvedAverage);
        user.totalPoints = user.positionPoints + user.ratingPoints + user.problemsPoints

    # Sort by points
    users.sort(key = lambda user : user.totalPoints, reverse = True)

    ranking = Ranking(users, totalUsers, ratingAverage, problemSolvedAverage)
    return ranking

def plotRanking(ranking):
    # Plot ranking
    headers = ["Id", "Name", "Handle", "Rating", "Problems", "PPosition", "PRatingP", "PProblems" ,"PTotal"]
    table = []
    for user in ranking.users:
        table.append([user.id, user.name, user.handle, user.rating, user.problems,  user.positionPoints, user.ratingPoints, user.problemsPoints, user.totalPoints])

    rankingTable = tabulate(table, headers=headers, tablefmt='orgtbl')
    print("Ranking completed!")
    print("")
    print("Total participants in the selection: {0}".format(ranking.totalUsers))
    print("Average rating of participants: {0}".format(ranking.ratingAverage))
    print("Average number of problems solved of participants: {0}".format(ranking.problemSolvedAverage))
    print("")
    print(rankingTable)


def main():
    filepath = sys.argv[1]

    data = readDataFromFile(filepath)
    users = getCompleteInfoByUser(data)

    totalParticipants = data.max_row - 1
    ratingWeight = 20
    problemsWeight = 30
    positionWeight = 50

    ranking = computeRanking(users, totalParticipants, ratingWeight, problemsWeight, positionWeight)
    plotRanking(ranking)

# Before run the script install dependencies
# pip3 install openpyxl
# pip3 install requests
# pip3 install json
# pip3 install tabulate

# Run command: python3 ranking.py /Users/xkleiber/Downloads/data_test.xlsx

main()


